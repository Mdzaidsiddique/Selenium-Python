import openpyxl

file = r"/Resources/14_data_driven_testing/sample2.xlsx"



workbook = openpyxl.load_workbook(file)
# sheet = workbook["Data"]
# or
sheet = workbook.active #get active (single) sheet


# for r in range(1, 5):       #4 rows
#     for c in range(1,4):    #3 col
#         sheet.cell(r,c).value = "Welcome"

# after that we have to save it

# workbook.save(file)

# writing multiple data
sheet.cell(1,2).value = "zaid"
sheet.cell(2,2).value = "Umair"
sheet.cell(3,4).value = 123
workbook.save(file)


