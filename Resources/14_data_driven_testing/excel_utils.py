import openpyxl
from openpyxl.styles import PatternFill


def get_row_ount(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_row

def get_column_ount(file, sheet_name):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet_name]
    return sheet.max_column

def read_data(file, sheet, row, column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    return sheet.cell(row, column).value

def write_excel(file, sheet, row, column, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    sheet.cell(row, column).value = data
    workbook.save(file)

def fill_green_colour(file, sheet, row, column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    color = PatternFill('solid', fgColor='00FF00') #green
    sheet.cell(row, column).fill = color
    workbook.save(file)

def fill_red_colour(file, sheet, row, column):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    color = PatternFill('solid', fgColor='ff0000') #red
    sheet.cell(row, column).fill = color
    workbook.save(file)