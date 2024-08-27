import openpyxl

file_path = r"C:\Users\mdzaids\Desktop\Selenium-Python\14_data_driven_testing\sample.xlsx"

# file -> workbook -> sheets -> rows -> cells

workbook = openpyxl.load_workbook(file_path)
sheet = workbook["Sheet1"]

rows = sheet.max_row
cols = sheet.max_column
print(rows,cols)

for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value)
    print()

